import os
import logging
import datetime
import pyodbc
from openpyxl import Workbook, load_workbook
import pdfplumber
from PyPDF2 import PdfReader, PdfWriter

# Configuração do log
logging.basicConfig(level=logging.INFO, filename="programa.log", format="%(asctime)s - %(levellevelname)s - %(message)s")

#iniciar a planilha que ficará com os registros que eu preciso baixar
def inicializar_planilha(caminho_planilha):
    diretorio = os.path.dirname(caminho_planilha)
    if not os.path.exists(diretorio):
        os.makedirs(diretorio)
    if not os.path.exists(caminho_planilha):
        wb = Workbook()
        ws = wb.active
        ws.append(['Centro de Custo', 'CPF', 'Nome', 'Conta','Agencia'])
        wb.save(caminho_planilha)
    else:
        wb = load_workbook(caminho_planilha)
    return wb
#Salva os dados
def salvar_dados_excel(caminho_planilha, dados):
    wb = load_workbook(caminho_planilha)
    ws = wb.active
    for dado in dados:
        ws.append(dado)
    wb.save(caminho_planilha)
#Remove linha quando o funcionário é encontrado
def remover_linha_excel(caminho_planilha, nome_funcionario):
    wb = load_workbook(caminho_planilha)
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Ignore header
        print(row[2].value)
        if row[2].value == nome_funcionario:
            ws.delete_rows(row[0].row, 1)
            break
    wb.save(caminho_planilha)
#Cria a conexão com o banco de dados
def conectar_banco_dados():
    try:
        conexao = pyodbc.connect(
            'DRIVER={driver};'
            'SERVER=server;'
            'DATABASE=db;'
            'UID=user;'
            'PWD=password;'
        )
        return conexao
    except pyodbc.Error as erro:
        logging.error(f"Erro ao conectar ao banco de dados: {erro}")
        return None
#Consulta que efetua a busca dos funcionários de acordo com a regra definida pela empresa
def buscar_funcionarios_ra_cic(conexao, cc):
    funcionarios_ra_cic = []
    if conexao:
        try:
            cursor = conexao.cursor()
            data_atual = datetime.datetime.now()
            primeiro_dia_mes_anterior = (data_atual.replace(day=1) - datetime.timedelta(days=1)).replace(day=1)
            ultimo_dia_mes_anterior = data_atual.replace(day=1) - datetime.timedelta(days=1)
            data_inicio = primeiro_dia_mes_anterior.strftime('%Y%m%d')
            data_fim = ultimo_dia_mes_anterior.strftime('%Y%m%d')
            consulta = f"""
                SELECT RA_CIC, RA_CTDEPSA , RA_BCDEPSA
                FROM SRA010 
                WHERE RA_CC ='{cc}' 
                AND (RA_DEMISSA = '' OR RA_DEMISSA >'{data_fim}')
                AND (RA_ADMISSA <= '{data_fim}')
                AND ((RA_MAT NOT IN (select RE_MATD FROM SRE010 WHERE (RE_DATA BETWEEN '{data_inicio}' AND '{data_fim}') AND (RE_CCD='{cc}') AND D_E_L_E_T_=''))
                OR (RA_MAT IN (select RE_MATD FROM SRE010 WHERE (RE_DATA BETWEEN '{data_inicio}' AND '{data_fim}') AND (RE_CCP='{cc}') AND D_E_L_E_T_='')))
            """
            cursor.execute(consulta)
            for row in cursor.fetchall():
                ra_cic = row[0]
                ra_ctdepsa = row[1]
                ra_bcdepsa = row[2]
                banco_codigo =ra_bcdepsa[:3]
                agencia = ra_bcdepsa[3:]
                ra_ctdepsa = ra_ctdepsa.strip()
                
                if banco_codigo == '001': #Layout BANCO DO BRASIL
                    # Formato 2: 000.000.000.888-5 (completar com zeros à esquerda para totalizar 11 dígitos)
                    num_conta= ra_ctdepsa.zfill(13)
                    num_conta = f"{num_conta[:3]}.{num_conta[3:6]}.{num_conta[6:9]}.{num_conta[9:12]}-{num_conta[12:]}"
                    agencia = agencia.zfill(5)
                    agencia_formatada = f"{agencia[0]}.{agencia[1:4]}-{agencia[-1]}"
                elif banco_codigo == '237': #Layout BRADESCO
                    # Formato 3: 000000000888-5 (completar com zeros à esquerda para totalizar 12 dígitos)
                    num_conta= ra_ctdepsa.zfill(12)
                    num_conta = f"{num_conta[:9]}-{num_conta[9:]}"
                    agencia = agencia.zfill(6)
                    agencia_formatada = f"{agencia[:5]}-{agencia[-1]}"
                elif banco_codigo == '341': #Layout ITAU
                    num_conta=  ra_ctdepsa.strip()[-6:]  # Pegar os últimos 6 caracteres
                    num_conta = f"{num_conta[:-1]} - {num_conta[-1]}"  # Formatar XXXXX - X  # Iterate over a copy of the list
                    agencia = agencia.zfill(4)
                    agencia_formatada = agencia
                
                funcionarios_ra_cic.append((ra_cic, num_conta,agencia_formatada))
            cursor.close()
        except pyodbc.Error as erro:
            logging.error(f"Erro ao buscar funcionários no banco de dados: {erro}")
    else:
        logging.error("Falha na conexão ao banco de dados.")
    return funcionarios_ra_cic
#Consulta que busca o nome completo do funcionário
def buscar_nome_completo(conexao, ra_cic):
    try:
        cursor = conexao.cursor()
        consulta = f"SELECT TRIM(RA_NOMECMP) FROM SRA010 WHERE RA_CIC = '{ra_cic}'"
        cursor.execute(consulta)
        resultado = cursor.fetchone()
        cursor.close()
        if resultado:
            return resultado[0]
        else:
            logging.error(f"Nome completo não encontrado para RA_CIC: {ra_cic}")
            return None
    except pyodbc.Error as erro:
        logging.error(f"Erro ao buscar o nome completo do funcionário: {ra_cic}, {erro}")
        return None
#Busca os funcionários do centro de custo adiciona o nome completo e salva os dados na planilha
def processar_centro_custo(conexao, cc, caminho_planilha):
    funcionarios = buscar_funcionarios_ra_cic(conexao, cc)
    dados = []
    for ra_cic, ra_ctdepsa,ra_bcdepsa in funcionarios:
        nome_completo = buscar_nome_completo(conexao, ra_cic)
        if nome_completo:
            dados.append([cc, ra_cic, nome_completo, ra_ctdepsa,ra_bcdepsa])
    salvar_dados_excel(caminho_planilha, dados)
    return dados  # Retorna os dados processados
#Função que extrai os dados do pdf
def extract_text_from_pdf(pdf_path):
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text += page.extract_text()
    return text
#Confere se existe o diretório, se não existir ele cria
def ensure_directory_exists(path):
    if not os.path.exists(path):
        os.makedirs(path)
#Divide o pdf a cada 50 páginas para otimizar a busca
def split_pdf(input_pdf_path, pages_per_split=50):
    """
    Divide um PDF em partes menores.
    """
    pdf = PdfReader(input_pdf_path)
    total_pages = len(pdf.pages)
    output_paths = []

    for i in range(0, total_pages, pages_per_split):
        pdf_writer = PdfWriter()
        for j in range(i, min(i + pages_per_split, total_pages)):
            pdf_writer.add_page(pdf.pages[j])
        
        output_path = f"{input_pdf_path[:-4]}_part{i//pages_per_split + 1}.pdf"
        with open(output_path, 'wb') as output_pdf:
            pdf_writer.write(output_pdf)
        
        output_paths.append(output_path)
    
    return output_paths

#Faz a busca de acordo com os dados dos funcionários nos pdfs de comprovante
def extract_employee_pages(pdf_path, nomes_func, output_dir, cc, caminho_planilha):
    ensure_directory_exists(output_dir)

    
    with pdfplumber.open(pdf_path) as pdf:
        num_pages = len(pdf.pages)
        for i in range(num_pages):
            text = pdf.pages[i].extract_text()
            for cc_cic, ra_ctdepsa, nome_completo,num_agencia in nomes_func[:]:
                ra_cic_formatado = f"{cc_cic[:3]}.{cc_cic[3:6]}.{cc_cic[6:9]}-{cc_cic[9:]}"
                if ((cc_cic in text) or (ra_cic_formatado in text)):
                    output_pdf_path = os.path.join(output_dir, f'{nome_completo}.pdf')
                    pdf_writer = PdfWriter()
                    # Se já existir o arquivo, alimenta ele
                    if os.path.exists(output_pdf_path):
                        existing_pdf_reader = PdfReader(output_pdf_path)
                        for page in existing_pdf_reader.pages:
                            pdf_writer.add_page(page)
                    pdf_writer.add_page(PdfReader(pdf_path).pages[i])
                    with open(output_pdf_path, 'wb') as out_pdf:
                        pdf_writer.write(out_pdf)
                    remover_linha_excel(caminho_planilha, nome_completo)
                    if (cc_cic, ra_ctdepsa, nome_completo) in nomes_func:
                        nomes_func.remove((cc_cic, ra_ctdepsa, nome_completo))
                    break
                elif ((ra_ctdepsa  in text) and (num_agencia in text)):
                    output_pdf_path = os.path.join(output_dir, f'{nome_completo}.pdf')
                    pdf_writer = PdfWriter()
                    if os.path.exists(output_pdf_path):
                        existing_pdf_reader = PdfReader(output_pdf_path)
                        for page in existing_pdf_reader.pages:
                            pdf_writer.add_page(page)
                    pdf_writer.add_page(PdfReader(pdf_path).pages[i])
                    with open(output_pdf_path, 'wb') as out_pdf:
                        pdf_writer.write(out_pdf)
                    remover_linha_excel(caminho_planilha, nome_completo)
                    if (cc_cic, ra_ctdepsa, nome_completo) in nomes_func:
                        nomes_func.remove((cc_cic, ra_ctdepsa, nome_completo))
                    break

#Depois faz uma busca no diretório e une todos os comprovantes dos funcionários em um só arquivo
def merge_pdfs(output_dir, final_output_path):
    pdf_writer = PdfWriter()
    for pdf_filename in os.listdir(output_dir):
        if pdf_filename.endswith('.pdf'):
            pdf_path = os.path.join(output_dir, pdf_filename)
            pdf_reader = PdfReader(pdf_path)
            for page in pdf_reader.pages:
                pdf_writer.add_page(page)
    
    with open(final_output_path, 'wb') as out_pdf:
        pdf_writer.write(out_pdf)
    print(f'Arquivo final gerado: {final_output_path}')

#A função principal que irá fazer a chamada das funções e definir os caminhos dos arquivos e os centros de custo
def main():
    ccs = ['cc1','cc2']  # Lista de centros de custo
    caminho_planilha = r"\\servidor\caminho\entradaCP.xlsx"
    input_pdf_dir = r"\\servidor\caminho\TESTE"  # Diretório onde estão os PDFs
    output_dir_base = r"\\servidor\caminho\TEMP"  # Diretório onde os PDFs temporários serão salvos
    final_output_dir = r"\\servidor\caminho\NOVOSPDF"  # Diretório onde os PDFs finais serão salvos

    # Inicializar a planilha
    inicializar_planilha(caminho_planilha)

    conexao = conectar_banco_dados()
    if conexao:
        for cc in ccs:
            funcionarios = processar_centro_custo(conexao, cc, caminho_planilha)
            nomes_func = [(dado[1], dado[3], dado[2],dado[4]) for dado in funcionarios]  # Obter RA_CIC, RA_CTDEPSA e Nome dos funcionários

            # Criar diretório para armazenar PDFs temporários do centro de custo
            cc_output_dir = os.path.join(output_dir_base, cc)
            ensure_directory_exists(cc_output_dir)

            for pdf_filename in os.listdir(input_pdf_dir):
                if pdf_filename.endswith('.pdf'):
                    pdf_path = os.path.join(input_pdf_dir, pdf_filename)
                    print(f'Processando {pdf_path}')

                    # Dividir o PDF em partes menores apenas uma vez
                    split_paths = split_pdf(pdf_path, pages_per_split=50)
                    for split_path in split_paths:
                        extract_employee_pages(split_path, nomes_func, cc_output_dir, cc, caminho_planilha)
                        os.remove(split_path)  # Remover arquivo temporário

            # Juntar todos os PDFs do centro de custo em um único PDF
            merge_pdfs(cc_output_dir, os.path.join(final_output_dir, f'{cc}.pdf'))
        
        conexao.close()
    else:
        logging.error("Não foi possível conectar ao banco de dados.")

if __name__ == "__main__":
    main()
